import os
import logging
import csv

from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s"
)
logger = logging.getLogger("ingestor")

INPUT_DIR = "/data/input"
OUTPUT_FILE = "/data/ingested.txt"
MAX_TABLE_ROWS = int(os.getenv("MAX_TABLE_ROWS", "200"))


def read_text_file(filepath):
    with open(filepath, "r", encoding="utf-8") as f:
        return f.read()


def read_csv_file(filepath):
    lines = []
    with open(filepath, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        for idx, row in enumerate(reader):
            if idx >= MAX_TABLE_ROWS:
                lines.append(f"... truncated after {MAX_TABLE_ROWS} rows ...")
                break
            cleaned = [str(cell).strip() for cell in row]
            lines.append(" | ".join(cleaned))
    return "\n".join(lines)


def read_excel_file(filepath):
    workbook = load_workbook(filepath, data_only=True, read_only=True)
    lines = []
    for sheet in workbook.worksheets:
        lines.append(f"[sheet] {sheet.title}")
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx >= MAX_TABLE_ROWS:
                lines.append(f"... truncated after {MAX_TABLE_ROWS} rows ...")
                break
            cleaned = ["" if cell is None else str(cell).strip() for cell in row]
            lines.append(" | ".join(cleaned))
        lines.append("")
    return "\n".join(lines).strip()


def extract_file_content(filepath):
    _, ext = os.path.splitext(filepath.lower())
    if ext in {".txt", ".md", ".log"}:
        return read_text_file(filepath)
    if ext == ".csv":
        return read_csv_file(filepath)
    if ext == ".xlsx":
        return read_excel_file(filepath)
    return None

def ingest():
    content = ""
    files_processed = 0
    for filename in sorted(os.listdir(INPUT_DIR)):
        filepath = os.path.join(INPUT_DIR, filename)
        if os.path.isfile(filepath):
            try:
                file_content = extract_file_content(filepath)
                if file_content is None:
                    logger.warning(f"Skipping unsupported file type: {filename}")
                    continue
                if not file_content.strip():
                    logger.warning(f"Skipping empty file: {filename}")
                    continue
                content += f"\n--- {filename} ---\n"
                content += file_content
                content += "\n"
                files_processed += 1
            except Exception as e:
                logger.error(f"Failed to read {filename}: {e}")

    if files_processed == 0:
        logger.warning("No input files found in /data/input/")

    with open(OUTPUT_FILE, "w", encoding="utf-8") as out:
        out.write(content)
    logger.info(f"Ingested {files_processed} files -> {OUTPUT_FILE}")

if __name__ == "__main__":
    ingest()
